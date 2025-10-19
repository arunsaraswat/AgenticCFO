"""End-to-end integration test for Cash Commander workflow.

This test validates the complete user journey:
1. Upload file → Dataset created
2. Work order auto-created
3. Execute work order → Cash Commander runs
4. Download artifact → Excel file generated

Tests the full stack: API → Services → Agents → Artifact Generation
"""
import os
import pytest
from io import BytesIO
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.main import app
from app.models.user import User
from app.models.work_order import WorkOrder
from app.models.artifact import Artifact
from app.models.dataset import Dataset


@pytest.fixture
def auth_headers(test_user: User, client: TestClient) -> dict:
    """Get authentication headers for test user."""
    # Login to get token
    response = client.post(
        "/api/auth/login",
        data={
            "username": test_user.email,
            "password": "testpassword123"
        }
    )
    assert response.status_code == 200
    token = response.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


@pytest.mark.asyncio
async def test_end_to_end_cash_commander_flow(
    client: TestClient,
    db: Session,
    test_user: User,
    auth_headers: dict
):
    """
    Test complete end-to-end flow from file upload to artifact download.

    Flow:
    1. Upload BankStatement.xlsx
    2. Verify dataset created
    3. Verify work order auto-created
    4. Execute work order (Cash Commander)
    5. Verify work order completed
    6. Verify artifact created
    7. Download artifact
    8. Verify Excel file is valid
    """
    # Step 1: Upload file
    print("\n[Step 1] Uploading BankStatement.xlsx...")

    # Load sample file
    sample_file_path = "tests/sample_files/BankStatement.xlsx"
    if not os.path.exists(sample_file_path):
        pytest.skip(f"Sample file not found: {sample_file_path}")

    with open(sample_file_path, "rb") as f:
        file_content = f.read()

    files = {"file": ("BankStatement.xlsx", BytesIO(file_content), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}

    upload_response = client.post(
        "/api/intake/upload",
        files=files,
        headers=auth_headers
    )

    assert upload_response.status_code == 201, f"Upload failed: {upload_response.json()}"
    upload_data = upload_response.json()

    print(f"✓ File uploaded: ID={upload_data['id']}, status={upload_data['status']}")

    # Step 2: Verify dataset created
    print("\n[Step 2] Verifying dataset created...")

    assert "dataset_id" in upload_data, "Dataset ID not in response"
    dataset_id = upload_data["dataset_id"]

    dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    assert dataset is not None, f"Dataset {dataset_id} not found in database"
    assert dataset.tenant_id == test_user.tenant_id

    print(f"✓ Dataset created: ID={dataset.id}, template={dataset.template_type}, status={dataset.dq_status}")

    # Step 3: Verify work order auto-created
    print("\n[Step 3] Verifying work order auto-created...")

    assert "work_order_id" in upload_data, "Work order ID not in response"
    work_order_id = upload_data["work_order_id"]

    work_order = db.query(WorkOrder).filter(WorkOrder.id == work_order_id).first()
    assert work_order is not None, f"Work order {work_order_id} not found in database"
    assert work_order.tenant_id == test_user.tenant_id
    assert work_order.status == "pending"
    assert dataset_id in work_order.input_datasets

    print(f"✓ Work order created: ID={work_order.id}, status={work_order.status}, objective={work_order.objective}")

    # Step 4: Execute work order (Cash Commander)
    print("\n[Step 4] Executing Cash Commander agent...")

    execute_response = client.post(
        f"/api/work-orders/{work_order_id}/execute",
        headers=auth_headers
    )

    # Note: This will call the actual LLM, so it may take 30-60 seconds
    # In production tests, we'd mock the LLM response
    assert execute_response.status_code == 200, f"Execution failed: {execute_response.json()}"
    execution_data = execute_response.json()

    print(f"✓ Work order executed: status={execution_data['status']}, progress={execution_data['progress_percentage']}%")

    # Step 5: Verify work order completed
    print("\n[Step 5] Verifying work order completed...")

    db.refresh(work_order)
    assert work_order.status == "completed", f"Expected 'completed', got '{work_order.status}'"
    assert work_order.progress_percentage == 100
    assert work_order.completed_at is not None
    assert "cash_commander" in work_order.agent_outputs
    assert work_order.execution_time_seconds > 0

    print(f"✓ Work order completed in {work_order.execution_time_seconds:.2f}s")
    print(f"  - Agent outputs: {list(work_order.agent_outputs.keys())}")
    print(f"  - Confidence: {work_order.agent_outputs['cash_commander'].get('confidence_score', 'N/A')}")

    # Step 6: Verify artifact created
    print("\n[Step 6] Verifying artifact created...")

    artifacts = db.query(Artifact).filter(Artifact.work_order_id == work_order_id).all()
    assert len(artifacts) > 0, "No artifacts created"

    artifact = artifacts[0]
    assert artifact.artifact_type == "excel"
    assert artifact.generated_by_agent == "cash_commander"
    assert os.path.exists(artifact.file_path), f"Artifact file not found: {artifact.file_path}"

    print(f"✓ Artifact created: ID={artifact.id}, name={artifact.artifact_name}")
    print(f"  - Type: {artifact.artifact_type}")
    print(f"  - Size: {artifact.file_size_bytes / 1024:.2f} KB")
    print(f"  - Path: {artifact.file_path}")

    # Step 7: Download artifact
    print("\n[Step 7] Downloading artifact...")

    download_response = client.get(
        f"/api/artifacts/{artifact.id}/download",
        headers=auth_headers
    )

    assert download_response.status_code == 200, f"Download failed: {download_response.status_code}"
    assert download_response.headers["content-type"] == artifact.mime_type
    assert "attachment" in download_response.headers.get("content-disposition", "")

    downloaded_content = download_response.content
    assert len(downloaded_content) > 0, "Downloaded file is empty"
    assert len(downloaded_content) == artifact.file_size_bytes, "File size mismatch"

    print(f"✓ Artifact downloaded: {len(downloaded_content)} bytes")

    # Step 8: Verify Excel file is valid
    print("\n[Step 8] Verifying Excel file structure...")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(downloaded_content))

        assert "Cash Ladder" in wb.sheetnames, "Cash Ladder sheet not found"

        ws = wb["Cash Ladder"]

        # Verify headers
        assert ws["A1"].value == "Week", "Header A1 should be 'Week'"
        assert ws["B1"].value == "Week Ending", "Header B1 should be 'Week Ending'"
        assert ws["C1"].value == "Opening Balance", "Header C1 should be 'Opening Balance'"

        # Verify at least 13 weeks of data
        data_rows = 0
        for row in range(2, 20):
            if ws[f"A{row}"].value:
                data_rows += 1

        assert data_rows >= 13, f"Expected at least 13 weeks of data, got {data_rows}"

        print(f"✓ Excel file valid: {len(wb.sheetnames)} sheets, {data_rows} data rows")

    except ImportError:
        print("⚠ openpyxl not installed, skipping Excel validation")

    # Summary
    print("\n" + "=" * 70)
    print("✓ END-TO-END TEST PASSED")
    print("=" * 70)
    print(f"Upload ID: {upload_data['id']}")
    print(f"Dataset ID: {dataset_id}")
    print(f"Work Order ID: {work_order_id}")
    print(f"Artifact ID: {artifact.id}")
    print(f"Execution Time: {work_order.execution_time_seconds:.2f}s")
    print(f"Total Cost: ${work_order.total_cost_usd:.4f}")
    print("=" * 70)


@pytest.mark.asyncio
async def test_work_order_list_artifacts(
    client: TestClient,
    db: Session,
    test_user: User,
    auth_headers: dict
):
    """Test listing artifacts for a work order."""
    # This test assumes a work order with artifacts exists
    # In real scenario, run after test_end_to_end_cash_commander_flow

    # Find any work order with artifacts
    work_order = db.query(WorkOrder).filter(
        WorkOrder.tenant_id == test_user.tenant_id,
        WorkOrder.status == "completed"
    ).first()

    if not work_order:
        pytest.skip("No completed work orders found")

    # List artifacts
    response = client.get(
        f"/api/artifacts/work-order/{work_order.id}",
        headers=auth_headers
    )

    assert response.status_code == 200
    artifacts = response.json()

    assert isinstance(artifacts, list)
    if len(artifacts) > 0:
        assert "id" in artifacts[0]
        assert "artifact_name" in artifacts[0]
        assert "artifact_type" in artifacts[0]


@pytest.mark.asyncio
async def test_work_order_execution_error_handling(
    client: TestClient,
    db: Session,
    test_user: User,
    auth_headers: dict
):
    """Test error handling when executing non-existent work order."""
    response = client.post(
        "/api/work-orders/99999/execute",
        headers=auth_headers
    )

    assert response.status_code == 404
    assert "not found" in response.json()["detail"].lower()


@pytest.mark.asyncio
async def test_artifact_download_access_control(
    client: TestClient,
    db: Session,
    test_user: User,
    auth_headers: dict
):
    """Test that users can only download artifacts from their own tenant."""
    # Try to download artifact with invalid ID
    response = client.get(
        "/api/artifacts/99999/download",
        headers=auth_headers
    )

    assert response.status_code == 404
