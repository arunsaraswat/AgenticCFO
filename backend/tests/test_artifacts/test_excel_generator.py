"""Tests for Excel artifact generator."""
import pytest
from pathlib import Path
from datetime import datetime, timedelta
import tempfile
import shutil

from openpyxl import load_workbook

from app.artifacts.excel_generator import ExcelGenerator, generate_cash_ladder


@pytest.fixture
def temp_output_dir():
    """Create temporary directory for test outputs."""
    temp_dir = tempfile.mkdtemp()
    yield temp_dir
    shutil.rmtree(temp_dir)


@pytest.fixture
def sample_forecast_data():
    """Generate sample forecast data for testing."""
    forecast = []
    base_date = datetime.now()
    current_balance = 1000000

    for week in range(1, 14):
        receipts = 250000 + (week * 5000)
        disbursements = 180000 + (week * 3000)
        ending_balance = current_balance + receipts - disbursements

        forecast.append({
            "week_number": week,
            "week_ending": (base_date + timedelta(weeks=week)).strftime("%Y-%m-%d"),
            "beginning_balance": current_balance,
            "cash_receipts": receipts,
            "cash_disbursements": disbursements,
            "ending_balance": ending_balance,
        })

        current_balance = ending_balance

    return forecast


class TestExcelGenerator:
    """Test Excel generator functionality."""

    def test_excel_generator_initialization(self, temp_output_dir):
        """Test ExcelGenerator initialization."""
        generator = ExcelGenerator(output_dir=temp_output_dir)
        assert generator.output_dir == Path(temp_output_dir)
        assert generator.output_dir.exists()

    def test_excel_generator_creates_directory(self):
        """Test that generator creates output directory if it doesn't exist."""
        with tempfile.TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir) / "nested" / "artifacts"
            generator = ExcelGenerator(output_dir=str(output_dir))
            assert output_dir.exists()

    def test_generate_cash_ladder_basic(self, temp_output_dir):
        """Test basic Cash Ladder generation."""
        generator = ExcelGenerator(output_dir=temp_output_dir)

        result = generator.generate_cash_ladder(
            current_cash=1000000.00,
            forecast_data=None,  # Will use sample data
            liquidity_warnings=None,
            recommendations=None,
        )

        # Verify result structure
        assert "file_path" in result
        assert "filename" in result
        assert "artifact_type" in result
        assert result["artifact_type"] == "excel"
        assert result["filename"].startswith("Cash_Ladder_")
        assert result["filename"].endswith(".xlsx")

        # Verify file exists
        assert Path(result["file_path"]).exists()

    def test_generate_cash_ladder_with_forecast_data(self, temp_output_dir, sample_forecast_data):
        """Test Cash Ladder generation with custom forecast data."""
        generator = ExcelGenerator(output_dir=temp_output_dir)

        liquidity_warnings = [
            "Cash balance falls below $500K in Week 8",
            "Minimum balance of $450K in Week 10",
        ]

        recommendations = [
            "Accelerate AR collections by offering 2% early payment discount",
            "Defer non-critical AP payments to Week 11-12",
            "Consider drawing $200K from credit line in Week 7",
        ]

        metadata = {
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "tenant_name": "Test Tenant Inc.",
            "user": "test_user@example.com",
        }

        result = generator.generate_cash_ladder(
            current_cash=1000000.00,
            forecast_data=sample_forecast_data,
            liquidity_warnings=liquidity_warnings,
            recommendations=recommendations,
            metadata=metadata,
        )

        # Verify file was created
        assert Path(result["file_path"]).exists()

        # Load and verify Excel content
        wb = load_workbook(result["file_path"])
        ws = wb.active

        # Verify title
        assert ws.title == "Cash Ladder"
        assert "13-Week Cash Forecast" in ws["A1"].value

        # Verify metadata
        assert "Test Tenant Inc." in ws["D2"].value

        # Verify headers (row 6)
        assert ws["A6"].value == "Week #"
        assert ws["B6"].value == "Week Ending"
        assert ws["C6"].value == "Beginning Balance"
        assert ws["D6"].value == "Cash Receipts"
        assert ws["E6"].value == "Cash Disbursements"
        assert ws["F6"].value == "Ending Balance"

        # Verify forecast data rows (rows 7-19 for 13 weeks)
        assert ws["A7"].value == 1  # Week 1
        assert ws["A19"].value == 13  # Week 13

        # Verify totals row exists
        assert ws["A21"].value == "TOTAL"

        wb.close()

    def test_generate_cash_ladder_file_size(self, temp_output_dir):
        """Test that generated file has reasonable size."""
        generator = ExcelGenerator(output_dir=temp_output_dir)

        result = generator.generate_cash_ladder(current_cash=1000000.00)

        # Verify size_bytes is returned and reasonable
        assert "size_bytes" in result
        assert result["size_bytes"] > 5000  # Should be at least 5KB
        assert result["size_bytes"] < 100000  # Should be less than 100KB

    def test_generate_cash_ladder_unique_filenames(self, temp_output_dir):
        """Test that each generation creates unique filename."""
        generator = ExcelGenerator(output_dir=temp_output_dir)

        result1 = generator.generate_cash_ladder(current_cash=1000000.00)
        result2 = generator.generate_cash_ladder(current_cash=1000000.00)

        assert result1["filename"] != result2["filename"]
        assert Path(result1["file_path"]).exists()
        assert Path(result2["file_path"]).exists()

    def test_generate_cash_ladder_with_warnings(self, temp_output_dir):
        """Test Cash Ladder with liquidity warnings."""
        generator = ExcelGenerator(output_dir=temp_output_dir)

        warnings = [
            "Cash balance falls below minimum threshold",
            "Covenant breach risk in Week 9",
        ]

        result = generator.generate_cash_ladder(
            current_cash=1000000.00,
            liquidity_warnings=warnings,
        )

        # Load and verify warnings are in the file
        wb = load_workbook(result["file_path"])
        ws = wb.active

        # Find warnings section (should contain "LIQUIDITY WARNINGS")
        found_warnings = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "LIQUIDITY WARNINGS" in str(cell.value):
                    found_warnings = True
                    break

        assert found_warnings
        wb.close()

    def test_generate_cash_ladder_with_recommendations(self, temp_output_dir):
        """Test Cash Ladder with recommendations."""
        generator = ExcelGenerator(output_dir=temp_output_dir)

        recommendations = [
            "Accelerate AR collections",
            "Defer AP payments",
        ]

        result = generator.generate_cash_ladder(
            current_cash=1000000.00,
            recommendations=recommendations,
        )

        # Load and verify recommendations are in the file
        wb = load_workbook(result["file_path"])
        ws = wb.active

        # Find recommendations section
        found_recommendations = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "RECOMMENDED ACTIONS" in str(cell.value):
                    found_recommendations = True
                    break

        assert found_recommendations
        wb.close()

    def test_generate_sample_forecast(self, temp_output_dir):
        """Test sample forecast generation."""
        generator = ExcelGenerator(output_dir=temp_output_dir)

        forecast = generator._generate_sample_forecast(starting_cash=1000000, num_weeks=13)

        assert len(forecast) == 13
        assert forecast[0]["week_number"] == 1
        assert forecast[12]["week_number"] == 13
        assert forecast[0]["beginning_balance"] == 1000000

        # Verify all required fields
        required_fields = [
            "week_number",
            "week_ending",
            "beginning_balance",
            "cash_receipts",
            "cash_disbursements",
            "ending_balance",
        ]
        for week_data in forecast:
            for field in required_fields:
                assert field in week_data

    def test_convenience_function(self, temp_output_dir):
        """Test the convenience function generate_cash_ladder()."""
        result = generate_cash_ladder(
            current_cash=1000000.00,
            output_dir=temp_output_dir,
        )

        assert "file_path" in result
        assert Path(result["file_path"]).exists()
        assert result["artifact_type"] == "excel"

    def test_generate_cash_ladder_with_zero_balance(self, temp_output_dir):
        """Test generation with zero cash balance."""
        generator = ExcelGenerator(output_dir=temp_output_dir)

        result = generator.generate_cash_ladder(current_cash=0.00)

        assert Path(result["file_path"]).exists()

        # Load and verify current cash is 0
        wb = load_workbook(result["file_path"])
        ws = wb.active

        # Find current cash position row
        found_zero = False
        for row in ws.iter_rows():
            if row[0].value == "Current Cash Position:":
                assert row[1].value == 0.00
                found_zero = True
                break

        assert found_zero
        wb.close()

    def test_generate_cash_ladder_with_negative_balance(self, temp_output_dir):
        """Test generation with negative cash balance."""
        generator = ExcelGenerator(output_dir=temp_output_dir)

        forecast_data = [{
            "week_number": 1,
            "week_ending": "2024-01-07",
            "beginning_balance": 100000,
            "cash_receipts": 50000,
            "cash_disbursements": 200000,
            "ending_balance": -50000,  # Negative!
        }]

        result = generator.generate_cash_ladder(
            current_cash=100000.00,
            forecast_data=forecast_data,
        )

        assert Path(result["file_path"]).exists()

        # Load and verify negative balance is formatted
        wb = load_workbook(result["file_path"])
        ws = wb.active

        # The ending balance should be -50000
        assert ws["F7"].value == -50000
        wb.close()


class TestExcelFormattingAndStructure:
    """Test Excel formatting and structure specifics."""

    def test_conditional_formatting_applied(self, temp_output_dir):
        """Test that conditional formatting is applied for low balances."""
        generator = ExcelGenerator(output_dir=temp_output_dir)

        # Create forecast with low balance
        forecast_data = [{
            "week_number": 1,
            "week_ending": "2024-01-07",
            "beginning_balance": 1000000,
            "cash_receipts": 100000,
            "cash_disbursements": 700000,
            "ending_balance": 400000,  # Below 500K threshold
        }]

        result = generator.generate_cash_ladder(
            current_cash=1000000.00,
            forecast_data=forecast_data,
        )

        wb = load_workbook(result["file_path"])
        ws = wb.active

        # Check that conditional formatting rules exist
        assert len(ws.conditional_formatting._cf_rules) > 0
        wb.close()

    def test_column_widths_set(self, temp_output_dir):
        """Test that column widths are properly set."""
        generator = ExcelGenerator(output_dir=temp_output_dir)
        result = generator.generate_cash_ladder(current_cash=1000000.00)

        wb = load_workbook(result["file_path"])
        ws = wb.active

        # Verify column widths are set (not default)
        assert ws.column_dimensions["A"].width > 0
        assert ws.column_dimensions["B"].width > 0
        assert ws.column_dimensions["C"].width > 0
        wb.close()

    def test_currency_formatting(self, temp_output_dir):
        """Test that currency values are properly formatted."""
        generator = ExcelGenerator(output_dir=temp_output_dir)
        result = generator.generate_cash_ladder(current_cash=1000000.00)

        wb = load_workbook(result["file_path"])
        ws = wb.active

        # Check current cash position formatting (row 4, column B)
        assert "$" in ws["B4"].number_format or "#,##0.00" in ws["B4"].number_format

        # Check forecast data currency formatting
        assert "$" in ws["C7"].number_format or "#,##0.00" in ws["C7"].number_format
        wb.close()
