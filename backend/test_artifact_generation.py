#!/usr/bin/env python
"""
Standalone test for Excel artifact generation.

This test verifies the artifact generation works independently of the agent infrastructure.
"""
import sys
from pathlib import Path
from datetime import datetime, timedelta
import tempfile

# Add backend to path
sys.path.insert(0, str(Path(__file__).parent))

from app.artifacts.excel_generator import generate_cash_ladder
from openpyxl import load_workbook


def test_excel_generation():
    """Test Excel generation with realistic cash forecast data."""
    print("\n" + "="*80)
    print("EXCEL ARTIFACT GENERATION TEST")
    print("="*80 + "\n")

    # Create temp directory for artifacts
    artifact_dir = tempfile.mkdtemp(prefix="agenticfo_test_")
    print(f"Artifact directory: {artifact_dir}\n")

    # Generate realistic forecast data
    print("Step 1: Generating 13-week cash forecast data...")
    forecast_data = []
    current_balance = 1245678.90  # Starting cash from bank statement
    base_date = datetime.now()

    for week in range(1, 14):
        # Simulate varying cash flows
        # Weeks 1-4: Healthy cash flows
        # Weeks 5-8: Declining receipts, higher disbursements
        # Weeks 9-13: Recovery

        if week <= 4:
            receipts = 180000 + (week * 2000)
            disbursements = 185000 + (week * 1500)
        elif week <= 8:
            receipts = 120000 - (week * 3000)  # Declining (creates low balance)
            disbursements = 195000 + (week * 5000)  # Increasing significantly
        else:
            receipts = 260000 + ((week - 8) * 4000)  # Recovery
            disbursements = 180000 + ((week - 8) * 1000)

        ending_balance = current_balance + receipts - disbursements

        forecast_data.append({
            "week_number": week,
            "week_ending": (base_date + timedelta(weeks=week)).strftime("%Y-%m-%d"),
            "beginning_balance": round(current_balance, 2),
            "cash_receipts": round(receipts, 2),
            "cash_disbursements": round(disbursements, 2),
            "ending_balance": round(ending_balance, 2),
        })

        current_balance = ending_balance

    print(f"✓ Generated {len(forecast_data)} weeks of forecast data")
    print(f"  Starting balance: ${forecast_data[0]['beginning_balance']:,.2f}")
    print(f"  Ending balance (Week 13): ${forecast_data[-1]['ending_balance']:,.2f}")

    # Find minimum balance week
    min_week = min(forecast_data, key=lambda x: x['ending_balance'])
    print(f"  Minimum balance: ${min_week['ending_balance']:,.2f} (Week {min_week['week_number']})")

    # Create warnings and recommendations
    print("\nStep 2: Generating liquidity warnings and recommendations...")
    liquidity_warnings = []
    recommendations = []

    if min_week['ending_balance'] < 500000:
        liquidity_warnings.append(
            f"Cash balance falls below minimum threshold ($500K) in Week {min_week['week_number']}"
        )
        liquidity_warnings.append(
            f"Minimum forecasted balance: ${min_week['ending_balance']:,.2f}"
        )
        recommendations.append("Accelerate AR collections by offering 2% early payment discount")
        recommendations.append("Defer non-critical vendor payments by 2 weeks")
        recommendations.append("Consider drawing from revolving credit line")

    if forecast_data[-1]['ending_balance'] < forecast_data[0]['beginning_balance']:
        liquidity_warnings.append("Net cash decrease over 13-week period")
        recommendations.append("Review operating expense budget for reduction opportunities")

    print(f"✓ Generated {len(liquidity_warnings)} warnings")
    print(f"✓ Generated {len(recommendations)} recommendations")

    # Generate Excel artifact
    print("\nStep 3: Generating Excel artifact...")
    metadata = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "tenant_name": "Demo Company Inc.",
        "agent": "cash_commander",
    }

    result = generate_cash_ladder(
        current_cash=forecast_data[0]['beginning_balance'],
        forecast_data=forecast_data,
        liquidity_warnings=liquidity_warnings,
        recommendations=recommendations,
        metadata=metadata,
        output_dir=artifact_dir,
    )

    print(f"✓ Excel file generated: {result['filename']}")
    print(f"  File size: {result['size_bytes']} bytes")
    print(f"  Artifact type: {result['artifact_type']}")

    # Verify Excel content
    print("\nStep 4: Verifying Excel content...")
    wb = load_workbook(result['file_path'])
    ws = wb.active

    # Verify structure
    tests_passed = 0
    tests_total = 0

    # Test 1: Sheet title
    tests_total += 1
    if ws.title == "Cash Ladder":
        print("  ✓ Sheet title correct")
        tests_passed += 1
    else:
        print(f"  ✗ Sheet title incorrect: {ws.title}")

    # Test 2: Main header
    tests_total += 1
    if ws["A1"].value and "13-Week Cash Forecast" in ws["A1"].value:
        print("  ✓ Main header present")
        tests_passed += 1
    else:
        print(f"  ✗ Main header missing or incorrect")

    # Test 3: Metadata
    tests_total += 1
    metadata_found = False
    for row in range(1, 10):
        if ws[f"D{row}"].value and "Demo Company Inc." in str(ws[f"D{row}"].value):
            metadata_found = True
            break
    if metadata_found:
        print("  ✓ Metadata present")
        tests_passed += 1
    else:
        print("  ✗ Metadata missing")

    # Test 4: Column headers
    tests_total += 1
    headers_found = False
    for row in range(1, 15):
        if ws[f"A{row}"].value == "Week #":
            headers_found = True
            header_row = row
            print(f"  ✓ Column headers found at row {row}")
            tests_passed += 1
            break
    if not headers_found:
        print("  ✗ Column headers missing")

    # Test 5: Forecast data rows
    tests_total += 1
    if headers_found:
        data_start_row = header_row + 1
        week1_value = ws[f"A{data_start_row}"].value
        week13_value = ws[f"A{data_start_row + 12}"].value
        if week1_value == 1 and week13_value == 13:
            print(f"  ✓ Forecast data rows present (Week 1-13)")
            tests_passed += 1
        else:
            print(f"  ✗ Forecast data rows incorrect: Week {week1_value} to {week13_value}")
    else:
        print("  ✗ Cannot verify forecast data (headers not found)")

    # Test 6: Current cash position
    tests_total += 1
    cash_position_found = False
    for row in range(1, 10):
        if ws[f"A{row}"].value == "Current Cash Position:":
            cash_value = ws[f"B{row}"].value
            if abs(cash_value - forecast_data[0]['beginning_balance']) < 0.01:
                print(f"  ✓ Current cash position correct: ${cash_value:,.2f}")
                tests_passed += 1
                cash_position_found = True
            else:
                print(f"  ✗ Current cash position incorrect: ${cash_value:,.2f}")
            break
    if not cash_position_found:
        print("  ✗ Current cash position not found")

    # Test 7: Warnings section
    tests_total += 1
    warnings_section_found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "LIQUIDITY WARNINGS" in str(cell.value):
                warnings_section_found = True
                break
    if warnings_section_found:
        print("  ✓ Liquidity warnings section present")
        tests_passed += 1
    else:
        print("  ✗ Liquidity warnings section missing")

    # Test 8: Recommendations section
    tests_total += 1
    recommendations_section_found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "RECOMMENDED ACTIONS" in str(cell.value):
                recommendations_section_found = True
                break
    if recommendations_section_found:
        print("  ✓ Recommendations section present")
        tests_passed += 1
    else:
        print("  ✗ Recommendations section missing")

    # Test 9: Key metrics section
    tests_total += 1
    key_metrics_found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "KEY METRICS" in str(cell.value):
                key_metrics_found = True
                break
    if key_metrics_found:
        print("  ✓ Key metrics section present")
        tests_passed += 1
    else:
        print("  ✗ Key metrics section missing")

    wb.close()

    # Print summary
    print("\n" + "="*80)
    if tests_passed == tests_total:
        print(f"ALL TESTS PASSED! ✓ ({tests_passed}/{tests_total})")
        print("="*80 + "\n")

        print("Summary:")
        print(f"  - Forecast weeks: 13")
        print(f"  - Starting cash: ${forecast_data[0]['beginning_balance']:,.2f}")
        print(f"  - Ending cash: ${forecast_data[-1]['ending_balance']:,.2f}")
        print(f"  - Minimum cash: ${min_week['ending_balance']:,.2f} (Week {min_week['week_number']})")
        print(f"  - Warnings: {len(liquidity_warnings)}")
        print(f"  - Recommendations: {len(recommendations)}")
        print(f"  - Excel file: {result['filename']}")
        print(f"  - File size: {result['size_bytes']:,} bytes")
        print(f"\n✓ Generated file can be opened at:")
        print(f"  {result['file_path']}")

        return True
    else:
        print(f"SOME TESTS FAILED ✗ ({tests_passed}/{tests_total} passed)")
        print("="*80 + "\n")
        return False


if __name__ == "__main__":
    try:
        success = test_excel_generation()
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"\n✗ TEST FAILED WITH EXCEPTION: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
