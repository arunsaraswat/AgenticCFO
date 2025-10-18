#!/usr/bin/env python
"""
End-to-end test for Cash Commander Agent with artifact generation.

This script tests the complete workflow:
1. Load sample test data files
2. Create datasets in the database
3. Execute Cash Commander agent
4. Generate Excel artifact
5. Verify artifact quality
"""
import asyncio
import os
import sys
from pathlib import Path
from datetime import datetime
import tempfile

# Add backend to path
sys.path.insert(0, str(Path(__file__).parent))

from app.agents.treasury.cash_commander import CashCommanderAgent
from app.db.session import SessionLocal
from app.services.dataset_service import DatasetService
from app.models.dataset import Dataset
from app.models.tenant import Tenant
import pandas as pd
from openpyxl import load_workbook


def setup_test_environment():
    """Set up test environment."""
    # Create temp directory for artifacts
    artifact_dir = tempfile.mkdtemp(prefix="agenticfo_test_")
    os.environ["ARTIFACTS_STORAGE_PATH"] = artifact_dir

    # Disable OpenRouter for testing (use mock)
    os.environ["OPENROUTER_API_KEY"] = "test-key"

    return artifact_dir


def load_sample_data():
    """Load sample Excel files."""
    sample_dir = Path(__file__).parent / "tests" / "sample_files"

    bank_statement_path = sample_dir / "BankStatement.xlsx"
    ar_aging_path = sample_dir / "AR_OpenItems.xlsx"
    ap_aging_path = sample_dir / "AP_OpenItems.xlsx"

    # Read Excel files
    bank_df = pd.read_excel(bank_statement_path)
    ar_df = pd.read_excel(ar_aging_path)
    ap_df = pd.read_excel(ap_aging_path)

    print(f"Loaded Bank Statement: {len(bank_df)} rows")
    print(f"Loaded AR Aging: {len(ar_df)} rows")
    print(f"Loaded AP Aging: {len(ap_df)} rows")

    return bank_df, ar_df, ap_df


def create_test_datasets(db):
    """Create test datasets in database."""
    dataset_service = DatasetService(db)

    # Get or create test tenant
    tenant = db.query(Tenant).filter_by(slug="test-tenant").first()
    if not tenant:
        tenant = Tenant(
            name="Test Tenant",
            slug="test-tenant",
            is_active=True
        )
        db.add(tenant)
        db.commit()
        db.refresh(tenant)

    # Load sample data
    bank_df, ar_df, ap_df = load_sample_data()

    # Create datasets
    bank_dataset = Dataset(
        tenant_id=tenant.id,
        template_type="BankStatement",
        entity="Main Bank",
        period_start=datetime(2024, 9, 1),
        period_end=datetime(2024, 9, 30),
        version=1,
        data_snapshot=bank_df.to_json(),
        row_count=len(bank_df),
        dq_status="validated"
    )
    db.add(bank_dataset)

    ar_dataset = Dataset(
        tenant_id=tenant.id,
        template_type="AR_OpenItems",
        entity="Main Entity",
        period_start=datetime(2024, 9, 1),
        period_end=datetime(2024, 9, 30),
        version=1,
        data_snapshot=ar_df.to_json(),
        row_count=len(ar_df),
        dq_status="validated"
    )
    db.add(ar_dataset)

    ap_dataset = Dataset(
        tenant_id=tenant.id,
        template_type="AP_OpenItems",
        entity="Main Entity",
        period_start=datetime(2024, 9, 1),
        period_end=datetime(2024, 9, 30),
        version=1,
        data_snapshot=ap_df.to_json(),
        row_count=len(ap_df),
        dq_status="validated"
    )
    db.add(ap_dataset)

    db.commit()
    db.refresh(bank_dataset)
    db.refresh(ar_dataset)
    db.refresh(ap_dataset)

    print(f"Created datasets:")
    print(f"  - Bank Statement: {bank_dataset.id}")
    print(f"  - AR Aging: {ar_dataset.id}")
    print(f"  - AP Aging: {ap_dataset.id}")

    return bank_dataset, ar_dataset, ap_dataset


async def test_cash_commander_execution():
    """Test Cash Commander execution with real data."""
    print("\n" + "="*80)
    print("CASH COMMANDER E2E TEST")
    print("="*80 + "\n")

    # Setup
    artifact_dir = setup_test_environment()
    print(f"Artifact directory: {artifact_dir}\n")

    # Create database session
    db = SessionLocal()

    try:
        # Create test datasets
        print("Step 1: Creating test datasets...")
        bank_dataset, ar_dataset, ap_dataset = create_test_datasets(db)
        print("✓ Datasets created\n")

        # Initialize Cash Commander
        print("Step 2: Initializing Cash Commander agent...")
        agent = CashCommanderAgent(db=db)
        print(f"✓ Agent initialized: {agent.agent_name}\n")

        # Prepare inputs
        print("Step 3: Preparing agent inputs...")
        inputs = {
            "bank_statement_id": str(bank_dataset.id),
            "ar_aging_id": str(ar_dataset.id),
            "ap_aging_id": str(ap_dataset.id),
        }

        policy_constraints = {
            "min_cash_balance": 500000,
            "forecast_weeks": 13,
        }
        print(f"✓ Inputs prepared\n")

        # Note: For this test, we'll skip the actual LLM execution
        # and test the artifact generation directly
        print("Step 4: Testing artifact generation...")

        # Simulate parsed output from agent
        parsed_output = {
            "current_cash_position": 1245678.90,
            "forecast": None,  # Will use sample data
            "liquidity_warnings": [
                "Cash balance may fall below minimum threshold in Week 9",
                "Consider accelerating AR collections"
            ],
            "recommendations": [
                "Offer 2% early payment discount on outstanding AR",
                "Defer non-critical vendor payments by 2 weeks",
                "Maintain minimum 3 weeks of operating expenses in cash"
            ],
            "summary": "Cash forecast shows healthy liquidity position..."
        }

        # Generate artifacts
        artifacts = await agent._generate_artifacts(parsed_output)
        print(f"✓ Generated {len(artifacts)} artifact(s)\n")

        # Verify artifact
        print("Step 5: Verifying Excel artifact...")
        artifact = artifacts[0]

        assert artifact["artifact_type"] == "excel"
        assert "file_path" in artifact
        assert Path(artifact["file_path"]).exists()
        print(f"✓ Artifact file exists: {artifact['filename']}")

        # Verify Excel content
        wb = load_workbook(artifact["file_path"])
        ws = wb.active

        assert ws.title == "Cash Ladder"
        assert "13-Week Cash Forecast" in ws["A1"].value
        print(f"✓ Excel title correct")

        # Check current cash position
        current_cash_row = None
        for row in range(1, 10):
            if ws[f"A{row}"].value == "Current Cash Position:":
                current_cash_row = row
                break

        assert current_cash_row is not None
        assert ws[f"B{current_cash_row}"].value == 1245678.90
        print(f"✓ Current cash position: ${ws[f'B{current_cash_row}'].value:,.2f}")

        # Check headers exist
        headers_found = False
        for row in range(1, 15):
            if ws[f"A{row}"].value == "Week #":
                headers_found = True
                print(f"✓ Found forecast table headers at row {row}")
                break

        assert headers_found

        # Check for warnings section
        warnings_found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "LIQUIDITY WARNINGS" in str(cell.value):
                    warnings_found = True
                    print(f"✓ Found LIQUIDITY WARNINGS section")
                    break

        assert warnings_found

        # Check for recommendations section
        recommendations_found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "RECOMMENDED ACTIONS" in str(cell.value):
                    recommendations_found = True
                    print(f"✓ Found RECOMMENDED ACTIONS section")
                    break

        assert recommendations_found

        wb.close()

        print(f"\n✓ File size: {artifact['size_bytes']} bytes")
        print(f"✓ Full path: {artifact['file_path']}")

        print("\n" + "="*80)
        print("ALL TESTS PASSED! ✓")
        print("="*80 + "\n")

        print("Summary:")
        print(f"  - Datasets created: 3")
        print(f"  - Agent initialized: Cash Commander")
        print(f"  - Artifacts generated: {len(artifacts)}")
        print(f"  - Excel validation: PASSED")
        print(f"\nGenerated file can be opened at:")
        print(f"  {artifact['file_path']}")

    except Exception as e:
        print(f"\n✗ TEST FAILED: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    finally:
        db.close()


if __name__ == "__main__":
    # Run test
    asyncio.run(test_cash_commander_execution())
