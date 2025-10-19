"""Excel artifact generator for financial reports and forecasts."""
from typing import Dict, List, Any, Optional
from datetime import datetime, timedelta
from pathlib import Path
import uuid

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


class ExcelGenerator:
    """Generate formatted Excel artifacts for financial reports."""

    def __init__(self, output_dir: str = "/tmp/artifacts"):
        """
        Initialize Excel generator.

        Args:
            output_dir: Directory to save generated Excel files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_cash_ladder(
        self,
        current_cash: float,
        forecast_data: Optional[List[Dict[str, Any]]] = None,
        liquidity_warnings: Optional[List[str]] = None,
        recommendations: Optional[List[str]] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        """
        Generate 13-week Cash Ladder Excel file.

        Args:
            current_cash: Current cash position
            forecast_data: List of weekly forecast dictionaries with:
                - week_number: int (1-13)
                - week_ending: date or str
                - beginning_balance: float
                - cash_receipts: float
                - cash_disbursements: float
                - ending_balance: float
            liquidity_warnings: List of warning messages
            recommendations: List of recommended actions
            metadata: Additional metadata (tenant, user, timestamp, etc.)

        Returns:
            Dict with artifact info:
                - file_path: str
                - filename: str
                - checksum_sha256: str (TODO: implement)
                - artifact_type: str
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Ladder"

        # Generate forecast data if not provided (for MVP/testing)
        if forecast_data is None:
            forecast_data = self._generate_sample_forecast(current_cash)

        # Write header section
        row = 1
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "13-Week Cash Forecast"
        ws[f"A{row}"].font = Font(size=16, bold=True, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 25

        # Metadata row
        row += 1
        if metadata:
            ws[f"A{row}"] = f"Generated: {metadata.get('generated_at', datetime.now().strftime('%Y-%m-%d %H:%M'))}"
            ws[f"D{row}"] = f"Tenant: {metadata.get('tenant_name', 'N/A')}"
        else:
            ws[f"A{row}"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        ws[f"A{row}"].font = Font(size=9, italic=True)
        ws[f"D{row}"].font = Font(size=9, italic=True)

        # Current cash position
        row += 2
        ws[f"A{row}"] = "Current Cash Position:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = current_cash
        ws[f"B{row}"].number_format = "$#,##0.00"
        ws[f"B{row}"].font = Font(bold=True, size=12)

        # Column headers
        row += 2
        headers = [
            ("Week #", 8),
            ("Week Ending", 12),
            ("Beginning Balance", 18),
            ("Cash Receipts", 15),
            ("Cash Disbursements", 18),
            ("Ending Balance", 18),
        ]

        header_row = row
        for col_idx, (header, width) in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                bottom=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Write forecast data
        row += 1
        first_data_row = row
        min_balance = float('inf')
        min_balance_week = None

        for week_data in forecast_data:
            week_num = week_data.get("week_number", 0)
            week_ending = week_data.get("week_ending", "")
            beg_balance = week_data.get("beginning_balance", 0)
            receipts = week_data.get("cash_receipts", 0)
            disbursements = week_data.get("cash_disbursements", 0)
            end_balance = week_data.get("ending_balance", 0)

            # Track minimum balance
            if end_balance < min_balance:
                min_balance = end_balance
                min_balance_week = week_num

            # Write row
            ws.cell(row=row, column=1, value=week_num)
            ws.cell(row=row, column=2, value=week_ending)
            ws.cell(row=row, column=3, value=beg_balance)
            ws.cell(row=row, column=4, value=receipts)
            ws.cell(row=row, column=5, value=disbursements)
            ws.cell(row=row, column=6, value=end_balance)

            # Format currency columns
            for col in [3, 4, 5, 6]:
                cell = ws.cell(row=row, column=col)
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")

            # Center align week number and date
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")

            # Add borders
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style="thin", color="D0D0D0"),
                    right=Side(style="thin", color="D0D0D0"),
                    top=Side(style="thin", color="D0D0D0"),
                    bottom=Side(style="thin", color="D0D0D0"),
                )

            row += 1

        last_data_row = row - 1

        # Add totals row (only if we have data)
        if forecast_data and len(forecast_data) > 0:
            row += 1
            ws.cell(row=row, column=1, value="TOTAL")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=4, value=f"=SUM(D{first_data_row}:D{last_data_row})")
            ws.cell(row=row, column=5, value=f"=SUM(E{first_data_row}:E{last_data_row})")

            # Format totals
            for col in [4, 5]:
                cell = ws.cell(row=row, column=col)
                cell.number_format = "$#,##0.00"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="right")
                cell.border = Border(
                    top=Side(style="double", color="000000"),
                    bottom=Side(style="double", color="000000"),
                )

            # Add conditional formatting for low balances (< $500k)
            low_balance_rule = CellIsRule(
                operator="lessThan",
                formula=["500000"],
                stopIfTrue=True,
                fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                font=Font(color="9C0006", bold=True),
            )
            ws.conditional_formatting.add(
                f"F{first_data_row}:F{last_data_row}",
                low_balance_rule,
            )

        # Liquidity warnings section
        if liquidity_warnings:
            row += 3
            ws.merge_cells(f"A{row}:F{row}")
            ws[f"A{row}"] = "LIQUIDITY WARNINGS"
            ws[f"A{row}"].font = Font(size=12, bold=True, color="FFFFFF")
            ws[f"A{row}"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            ws[f"A{row}"].alignment = Alignment(horizontal="center")

            row += 1
            for warning in liquidity_warnings:
                ws.merge_cells(f"A{row}:F{row}")
                ws[f"A{row}"] = f"• {warning}"
                ws[f"A{row}"].font = Font(color="C00000")
                ws[f"A{row}"].alignment = Alignment(wrap_text=True)
                row += 1

        # Recommendations section
        if recommendations:
            row += 2
            ws.merge_cells(f"A{row}:F{row}")
            ws[f"A{row}"] = "RECOMMENDED ACTIONS"
            ws[f"A{row}"].font = Font(size=12, bold=True, color="FFFFFF")
            ws[f"A{row}"].fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
            ws[f"A{row}"].alignment = Alignment(horizontal="center")

            row += 1
            for rec in recommendations:
                ws.merge_cells(f"A{row}:F{row}")
                ws[f"A{row}"] = f"• {rec}"
                ws[f"A{row}"].font = Font(color="375623")
                ws[f"A{row}"].alignment = Alignment(wrap_text=True)
                row += 1

        # Key metrics summary
        row += 2
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "KEY METRICS"
        ws[f"A{row}"].font = Font(size=12, bold=True, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")
        ws[f"A{row}"].alignment = Alignment(horizontal="center")

        row += 1
        ws[f"A{row}"] = "Minimum Forecasted Balance:"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = min_balance
        ws[f"B{row}"].number_format = "$#,##0.00"
        ws[f"B{row}"].font = Font(bold=True)
        ws[f"D{row}"] = f"(Week {min_balance_week})"

        # Save file with UUID
        filename = f"Cash_Ladder_{uuid.uuid4().hex[:8]}.xlsx"
        file_path = self.output_dir / filename

        wb.save(str(file_path))

        return {
            "file_path": str(file_path),
            "filename": filename,
            "artifact_type": "excel",
            "description": "13-week cash forecast ladder",
            "size_bytes": file_path.stat().st_size,
        }

    def _generate_sample_forecast(self, starting_cash: float, num_weeks: int = 13) -> List[Dict[str, Any]]:
        """
        Generate sample forecast data for testing.

        Args:
            starting_cash: Starting cash balance
            num_weeks: Number of weeks to forecast

        Returns:
            List of weekly forecast dictionaries
        """
        forecast = []
        current_balance = starting_cash
        base_date = datetime.now()

        for week in range(1, num_weeks + 1):
            # Simple heuristic: receipts vary, disbursements are steady
            receipts = 250000 + (week * 5000)  # Increasing receipts
            disbursements = 180000 + (week * 3000)  # Increasing disbursements
            ending_balance = current_balance + receipts - disbursements

            week_ending = base_date + timedelta(weeks=week)

            forecast.append({
                "week_number": week,
                "week_ending": week_ending.strftime("%Y-%m-%d"),
                "beginning_balance": current_balance,
                "cash_receipts": receipts,
                "cash_disbursements": disbursements,
                "ending_balance": ending_balance,
            })

            current_balance = ending_balance

        return forecast


def generate_cash_ladder(
    current_cash: float,
    forecast_data: Optional[List[Dict[str, Any]]] = None,
    liquidity_warnings: Optional[List[str]] = None,
    recommendations: Optional[List[str]] = None,
    metadata: Optional[Dict[str, Any]] = None,
    output_dir: str = "/tmp/artifacts",
) -> Dict[str, Any]:
    """
    Convenience function to generate Cash Ladder Excel file.

    Args:
        current_cash: Current cash position
        forecast_data: Weekly forecast data
        liquidity_warnings: List of warnings
        recommendations: List of recommendations
        metadata: Additional metadata
        output_dir: Output directory for Excel file

    Returns:
        Dict with artifact information
    """
    generator = ExcelGenerator(output_dir=output_dir)
    return generator.generate_cash_ladder(
        current_cash=current_cash,
        forecast_data=forecast_data,
        liquidity_warnings=liquidity_warnings,
        recommendations=recommendations,
        metadata=metadata,
    )
